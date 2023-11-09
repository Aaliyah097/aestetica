import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill
import bs4
import pandas

from src.staff.entities.users.staff import Staff
from src.staff.entities.users.doctor import Doctor
from src.staff.entities.users.seller import Seller
from src.staff.entities.users.assistant import Assistant
from src.staff.entities.department import Department
from src.salary.repositories.salary_repository import SalaryRepository
from src.salary.entities.salary_grid import SalaryGrid
from src.staff.repositories.filials_repository import FilialsRepository
from src.staff.repositories.departments_repository import DepartmentsRepository


class SalaryManagementService:
    def __init__(self):
        self.salary_repo: SalaryRepository = SalaryRepository()

    def modify_salary(self, salary_id: int, fix: 0, grid: list[dict]) -> None:
        if not fix or fix < 0:
            fix = 0

        if not grid:
            grid = []

        salary_grid = [
            SalaryGrid(
                limit=g['limit'],
                percent=g['percent']
            )
            for g in grid
        ]

        self.salary_repo.modify_salary(salary_id=salary_id, fix=fix, grid=salary_grid)

    def create_salary_for_staff(self, staff: Staff) -> None:
        departments = DepartmentsRepository.get_all()
        if isinstance(staff, Doctor):
            for filial in FilialsRepository.get_all():
                for department in filter(lambda d: d.name != 'Прочее', departments):
                    salary = self.salary_repo.create_salary(
                        staff=staff,
                        department=department,
                        fix=0,
                        filial=filial
                    )

                    if not salary:
                        continue

                    default_grids = {
                        1_500_000: 20,
                        2_500_000: 25,
                        100_000_000: 30
                    }

                    self.salary_repo.create_grid(
                        salary=salary,
                        grid=[
                            SalaryGrid(
                                limit=key,
                                percent=value
                            )
                            for key, value in default_grids.items()
                        ]
                    )
        elif isinstance(staff, Seller):
            for filial in FilialsRepository.get_all():
                salary = self.salary_repo.create_salary(
                    staff=staff,
                    department=Department(name='Прочее'),
                    fix=25000,
                    filial=filial
                )

                if not salary:
                    continue

                default_grids = {
                    5_000_000: 2,
                    10_000_000: 2.5,
                    100_000_000: 3
                }

                self.salary_repo.create_grid(
                    salary=salary,
                    grid=[
                        SalaryGrid(
                            limit=key,
                            percent=value
                        )
                        for key, value in default_grids.items()
                    ]
                )
        else:
            for filial in FilialsRepository.get_all():
                self.salary_repo.create_salary(
                    staff=staff,
                    department=Department(name='Прочее'),
                    fix=5000 if isinstance(staff, Assistant) else 0,
                    filial=filial
                )

    # TODO создать отдельный класс генератора отчетов в excel
    @staticmethod
    def export_salary(table_html: str, filial: str = None, date_begin: datetime.date = None, date_end: datetime.date = None) -> None:
        df = pandas.read_html(table_html)[0]

        file_name = f"static/{filial}_{str(date_begin)}_{str(date_end)}.xlsx"

        writer = pandas.ExcelWriter(file_name, engine='openpyxl')

        df.to_excel(writer, sheet_name=f"{str(date_begin)}-{str(date_end)}", index=False)
        sheet = writer.sheets[f"{str(date_begin)}-{str(date_end)}"]

        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row, col).value = sheet.cell(row, col).value.replace(" ₽", "").replace("₽", "")

        for idx, column in enumerate(sheet.columns):
            if idx == 3:
                continue
            max_length = 0
            column_name = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)

            sheet.column_dimensions[column_name].width = adjusted_width

        bold_font = Font(bold=True)
        blue = PatternFill(start_color='CFE2FF', end_color='CFE2FF', fill_type='solid')
        green = PatternFill(start_color='A7C9A7', end_color='A7C9A7', fill_type='solid')
        dark_green = PatternFill(start_color='839D83', end_color='839D83', fill_type='solid')
        orange = PatternFill(start_color='FAE08B', end_color='FAE08B', fill_type='solid')

        start_counter, end_counter = None, None
        min_row = 2

        for row_idx, row in enumerate(sheet.iter_rows(min_row=min_row, values_only=True)):
            rows_idx = row_idx + min_row
            if row[0] in ['Итого', "Клиент", 'Дата', 'Всего'] or row[2] == 'Руб.':
                for col_idx in range(1, sheet.max_column + 1):
                    sheet.cell(rows_idx, col_idx).font = bold_font
                    if row[0] in ['Клиент', 'Дата']:
                        sheet.cell(rows_idx, col_idx).fill = blue
                    elif row[0] == 'Итого':
                        sheet.cell(rows_idx, col_idx).fill = green
                    elif row[2] == 'Руб.':
                        sheet.cell(rows_idx, col_idx).fill = orange
                    elif row[0] == 'Всего':
                        sheet.cell(rows_idx, col_idx).fill = dark_green
            else:
                if sheet.cell(rows_idx - 1, 1).value in ['Клиент', 'Дата']:
                    if not start_counter:
                        start_counter = rows_idx - 1
                if sheet.cell(rows_idx + 1, 1).value == 'Итого':
                    if not end_counter:
                        end_counter = rows_idx + 1

            if start_counter and end_counter:
                for i in range(start_counter, end_counter + 1):
                    sheet.row_dimensions[i].outline_level = 1
                    sheet.row_dimensions[i].outline_below = True
                    sheet.row_dimensions[i].hidden = True
                start_counter, end_counter = None, None

            for i in range(1, sheet.max_column + 1):
                if "Unnamed" in sheet.cell(1, i).value:
                    sheet.cell(1, i).value = ""

            sheet.freeze_panes = "A2"

        writer.close()
